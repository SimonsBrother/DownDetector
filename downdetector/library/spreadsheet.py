""" Contains functions for reading data from the site spreadsheet """
import os

import openpyxl.worksheet.worksheet as pyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles.fills import PatternFill
from openpyxl import load_workbook

from downdetector.library.classes import Server
from downdetector.library.pinging import validateIP
# todo: IP validation

mac_excel_path = r"/Volumes/Public/Grade 3 Security Documents/BVH/BVH CCTV Support Checks/DownDetectorServers.xlsx"
windows_excel_path = r"X:\Grade 3 Security Documents\BVH\BVH CCTV Support Checks\DownDetectorServers.xlsx"

ONLINE = "ONLINE"
OFFLINE = "OFFLINE"


def getTestExcelPath():
    """ Returns the valid path for testing from within this python file """
    return windows_excel_path if os.name == "nt" else mac_excel_path


def spreadsheetIsPresent(path):
    """ Check the file exists """
    return (os.path.isfile(path) and str(path).endswith(".xlsx")) or path == ""


def getServerWorksheet(workbook: Workbook):
    """ Loads the worksheet from the Excel file from the path """
    return workbook["Servers"]


def getServers(path):
    """
    Gets all the servers from worksheet provided as a list of Server objects
    :param path: The path to the spreadsheet
    :return: A list of all the servers
    """

    workbook = load_workbook(filename=path)
    worksheet = workbook["Servers"]

    current_site_name = ""
    server_list = []

    # Work down the spreadsheet, changing the current site name as they are encountered, creating server objects with
    #  the IP addresses and statuses.
    # For each row
    for row in worksheet.iter_rows(min_row=2, max_col=3):

        # Change the site name "buffer" if a new name is encountered
        if row[0].value is not None:
            current_site_name = str(row[0].value)

        # If an IP is available
        if row[1].value is not None:
            ip = str(row[1].value).replace(" ", "")  # Get the IP in string form without spaces

            status = str(row[2].value)

            server_list.append(Server(current_site_name, ip, status == ONLINE, row[2].coordinate))

    workbook.close()
    return server_list


def tweakWorksheet(path):
    """ Removes the statuses of blank rows, and sets any status that isn't 'Online' to 'Offline' """

    workbook = load_workbook(filename=path)
    worksheet = getServerWorksheet(workbook)

    # For each row
    for row in worksheet.iter_rows(min_row=2, max_col=3):

        # If there is an IP
        if row[1].value is not None:
            status = str(row[2].value)

            # Check if its status is valid
            if status not in (ONLINE, OFFLINE):
                # If invalid, set its state to OFFLINE
                setCellStatus(row[2].coordinate, OFFLINE, worksheet)
            else:
                setCellStatus(row[2].coordinate, status, worksheet)
        else:
            setCellStatus(row[2].coordinate, None, worksheet)

    workbook.save(path)
    workbook.close()


def setCellStatus(cell: str, status, worksheet: pyxl.Worksheet):
    """
    Sets the status of a cell (for colouring later) using the path to the spreadsheet. Must be saved externally.
    :param cell: The coordinates of the cell to set the status of e.g. C3
    :param status: None, ONLINE, or OFFLINE
    :param worksheet: The worksheet of the server spreadsheet
    """

    colour = "00333333"
    if status == ONLINE or status is True:
        worksheet[cell].value = ONLINE
        colour = "0000FF00"

    elif status == OFFLINE or status is False:
        worksheet[cell].value = OFFLINE
        colour = "00FF0000"

    elif status is None:
        worksheet[cell].value = None

    else:
        raise ValueError("Invalid status")

    worksheet[cell].fill = PatternFill(fill_type="solid", bgColor=colour, fgColor=colour)


if __name__ == "__main__":
    tweakWorksheet(getTestExcelPath())
    servers = getServers(getTestExcelPath())
    for s in servers:
        print(s)
